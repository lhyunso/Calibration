"""
XLSX calibration report generator.

Sheet layout
────────────
  Cover_Summary  : metadata + 2-1 / 2-2 all-channel results
  CH01 … CH16    : per-channel detail (decimals / voltages / resistances / charts)
  3W_Reference   : 3-wire reference voltage table for the active sensor type
  Raw_Data       : decimal AVG/MIN/MAX per resistance step × channel
"""
import io
import os
from datetime import datetime
from typing import Dict, List, Optional, Tuple

import matplotlib
matplotlib.use("Agg")
import matplotlib.pyplot as plt
import matplotlib.ticker as mticker
import numpy as np

import openpyxl
from openpyxl import Workbook
from openpyxl.styles import (
    PatternFill, Font, Alignment, Border, Side, numbers
)
from openpyxl.utils import get_column_letter
from openpyxl.drawing.image import Image as XLImage

from processing.calibration import ChannelCalibration
from sensors.base import SensorConfig
from reference.three_wire import build_reference_table, STANDARD_GAINS


# ── Colour palette (hex, no #) ────────────────────────────────────────────────
C_NAVY   = "1F497D"
C_LBLUE  = "D6E4F0"
C_ALTROW = "F2F7FB"
C_PASS   = "0070C0"
C_FAIL   = "C00000"
C_WHITE  = "FFFFFF"
C_BLACK  = "000000"
C_GREY   = "A0A0A0"

# ── Style factories ───────────────────────────────────────────────────────────

def _fill(hex_color: str) -> PatternFill:
    return PatternFill("solid", fgColor=hex_color)


def _font(bold=False, color=C_BLACK, size=9, name="맑은 고딕") -> Font:
    return Font(bold=bold, color=color, size=size, name=name)


def _align(h="center", v="center", wrap=False) -> Alignment:
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)


def _border(style="thin", color=C_GREY) -> Border:
    s = Side(border_style=style, color=color)
    return Border(left=s, right=s, top=s, bottom=s)


def _thin_border() -> Border:
    return _border("thin", C_GREY)


# ── Cell helpers ──────────────────────────────────────────────────────────────

def _hdr(ws, row, col, text, size=9, colspan=1, rowspan=1):
    """Navy-background white-text header cell."""
    cell = ws.cell(row=row, column=col, value=text)
    cell.fill   = _fill(C_NAVY)
    cell.font   = _font(bold=True, color=C_WHITE, size=size)
    cell.alignment = _align()
    cell.border = _thin_border()
    if colspan > 1 or rowspan > 1:
        ws.merge_cells(
            start_row=row, start_column=col,
            end_row=row + rowspan - 1, end_column=col + colspan - 1,
        )
    return cell


def _sec(ws, row, col, text, colspan=1, rowspan=1):
    """Light-blue section label cell."""
    cell = ws.cell(row=row, column=col, value=text)
    cell.fill   = _fill(C_LBLUE)
    cell.font   = _font(bold=True, size=9)
    cell.alignment = _align(h="left")
    cell.border = _thin_border()
    if colspan > 1 or rowspan > 1:
        ws.merge_cells(
            start_row=row, start_column=col,
            end_row=row + rowspan - 1, end_column=col + colspan - 1,
        )
    return cell


def _lbl(ws, row, col, text, bold=False, colspan=1):
    """Plain label cell."""
    cell = ws.cell(row=row, column=col, value=text)
    cell.font   = _font(bold=bold, size=9)
    cell.alignment = _align(h="left")
    cell.border = _thin_border()
    if colspan > 1:
        ws.merge_cells(
            start_row=row, start_column=col,
            end_row=row, end_column=col + colspan - 1,
        )
    return cell


def _val(ws, row, col, value, fmt=None, color=C_BLACK, h="center", colspan=1):
    """Data value cell."""
    cell = ws.cell(row=row, column=col, value=value)
    cell.font      = _font(color=color, size=9)
    cell.alignment = _align(h=h)
    cell.border    = _thin_border()
    if fmt:
        cell.number_format = fmt
    if colspan > 1:
        ws.merge_cells(
            start_row=row, start_column=col,
            end_row=row, end_column=col + colspan - 1,
        )
    return cell


def _blank_row(ws, row, n_cols):
    """Write a blank merged spacer row."""
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=n_cols)
    ws.row_dimensions[row].height = 6


def _fmt_f(v, digits=6) -> str:
    if v is None:
        return "-"
    return f"{v:.{digits}f}"


def _pass_color(dev: float, tol: float) -> str:
    return C_PASS if abs(dev) <= tol else C_FAIL


# ── Chart generators ──────────────────────────────────────────────────────────

def _chart_gain(cal: ChannelCalibration) -> io.BytesIO:
    """Scatter + regression line: measured voltage vs reference voltage."""
    rs     = sorted(cal.voltages_avg.keys())
    v_meas = [cal.voltages_avg[r] for r in rs]
    v_ref  = [cal.voltage_ref[r]  for r in rs]

    fig, ax = plt.subplots(figsize=(4.8, 3.0))
    ax.scatter(v_ref, v_meas, color=f"#{C_NAVY}", s=50, zorder=5, label="Measured")
    if len(v_ref) >= 2:
        coeffs = np.polyfit(v_ref, v_meas, 1)
        x_line = np.linspace(min(v_ref) * 1.05, max(v_ref) * 1.05, 100)
        ax.plot(x_line, np.polyval(coeffs, x_line),
                color=f"#{C_FAIL}", lw=1.5, label=f"Gain={cal.gain:.6f}")
    ax.set_xlabel("Reference Voltage (V)", fontsize=7)
    ax.set_ylabel("Measured Voltage (V)", fontsize=7)
    ax.set_title(f"Gain — {cal.channel}", fontsize=8, fontweight="bold")
    ax.legend(fontsize=6)
    ax.grid(True, ls="--", alpha=0.4)
    ax.tick_params(labelsize=6)
    fig.tight_layout()
    buf = io.BytesIO()
    fig.savefig(buf, format="png", dpi=130)
    plt.close(fig)
    buf.seek(0)
    return buf


def _chart_deviation(cal: ChannelCalibration, resistances: List[float]) -> io.BytesIO:
    """Bar chart: deviation by resistance for both offset methods."""
    rs      = sorted(r for r in resistances if r in cal.dev_final_100)
    dev_100  = [cal.dev_final_100.get(r, 0) for r in rs]
    dev_mean = [cal.dev_final_mean.get(r, 0) for r in rs]
    tol      = cal.tolerance_max_100.get(rs[0], 0.385) if rs else 0.385

    x, w = np.arange(len(rs)), 0.35
    fig, ax = plt.subplots(figsize=(4.8, 3.0))
    ax.bar(x - w/2, dev_100,  w, color=f"#{C_NAVY}",  alpha=0.8, label="2-1 (100Ω)")
    ax.bar(x + w/2, dev_mean, w, color="#ED7D31", alpha=0.8, label="2-2 (Mean)")
    ax.axhline( tol, color="red",  ls="--", lw=1.0, label=f"+{tol}Ω")
    ax.axhline(-tol, color="blue", ls="--", lw=1.0, label=f"-{tol}Ω")
    ax.axhline(0,    color="black", ls="-", lw=0.5)
    ax.set_xticks(x)
    ax.set_xticklabels([f"{int(r)}Ω" for r in rs], fontsize=6)
    ax.set_ylabel("Deviation (Ω)", fontsize=7)
    ax.set_title(f"Deviation — {cal.channel}", fontsize=8, fontweight="bold")
    ax.legend(fontsize=6, loc="upper right")
    ax.grid(True, axis="y", ls="--", alpha=0.4)
    ax.tick_params(labelsize=6)
    fig.tight_layout()
    buf = io.BytesIO()
    fig.savefig(buf, format="png", dpi=130)
    plt.close(fig)
    buf.seek(0)
    return buf


# ── Sheet builders ────────────────────────────────────────────────────────────

class CalibrationXlsxWriter:
    """Generates a multi-sheet calibration XLSX report."""

    # Resistance column offset in detail sheets
    _RES_COL_START = 3   # column C onward for resistance values

    def __init__(
        self,
        sensor: SensorConfig,
        calibrations: Dict[str, ChannelCalibration],
        metadata: Optional[Dict[str, str]] = None,
    ):
        self.sensor       = sensor
        self.cals         = calibrations
        self.channels     = sorted(calibrations.keys())
        self.meta         = metadata or {}
        self.resistances: List[float] = []
        if calibrations:
            first = next(iter(calibrations.values()))
            self.resistances = sorted(first.voltages_avg.keys())

    # ── Cover / Summary sheet ─────────────────────────────────────────────────

    def _build_cover(self, wb: Workbook):
        ws = wb.active
        ws.title = "Cover_Summary"
        ws.sheet_view.showGridLines = False

        N = 10  # working column width

        # ── Title block ──────────────────────────────────────────────────────
        ws.row_dimensions[1].height = 36
        ws.merge_cells("A1:J1")
        title_cell = ws["A1"]
        title_cell.value     = "CALIBRATION REPORT"
        title_cell.font      = Font(bold=True, size=20, color=C_NAVY, name="맑은 고딕")
        title_cell.alignment = _align(h="center", v="center")
        title_cell.fill      = _fill(C_ALTROW)

        ws.row_dimensions[2].height = 16
        ws.merge_cells("A2:G2")
        doc_no = self.meta.get("doc_number", f"CAL-{datetime.now().strftime('%Y')}-0001")
        ws["A2"].value     = f"문서번호 (Doc. No.): {doc_no}"
        ws["A2"].font      = _font(size=9)
        ws["A2"].alignment = _align(h="left")
        ws.merge_cells("H2:J2")
        ws["H2"].value     = f"Rev: {self.meta.get('revision', '00')}"
        ws["H2"].font      = _font(size=9)
        ws["H2"].alignment = _align(h="right")

        # ── Section 1: Module Information ────────────────────────────────────
        r = 4
        _blank_row(ws, r - 1, N); r -= 1; r += 1
        _sec(ws, r, 1, "1. 모듈 정보 (Module Information)", colspan=N); r += 1

        pairs = [
            ("모듈명 (Module Name)", "module_name", "모델명 (Model No.)",       "model"),
            ("시리얼 번호 (S/N)",    "serial",      "제조사 (Manufacturer)",     "manufacturer"),
            ("FW / SW 버전",         "fw_version",  "",                          ""),
        ]
        for lbl1, key1, lbl2, key2 in pairs:
            _lbl(ws, r, 1, lbl1, bold=True)
            _val(ws, r, 2, self.meta.get(key1, ""), h="left", colspan=3)
            if lbl2:
                _lbl(ws, r, 6, lbl2, bold=True)
                _val(ws, r, 7, self.meta.get(key2, ""), h="left", colspan=4)
            r += 1

        # ── Section 2: Calibration Conditions ────────────────────────────────
        _blank_row(ws, r, N); r += 1
        _sec(ws, r, 1, "2. 교정 조건 (Calibration Conditions)", colspan=N); r += 1

        cond_pairs = [
            ("교정 일자 (Date)",    "date",        "교정 장소 (Location)",   "location"),
            ("담당자 (Technician)", "operator",    "온도 / 습도",             "temp_humidity"),
        ]
        for lbl1, key1, lbl2, key2 in cond_pairs:
            _lbl(ws, r, 1, lbl1, bold=True)
            _val(ws, r, 2, self.meta.get(key1, ""), h="left", colspan=3)
            _lbl(ws, r, 6, lbl2, bold=True)
            _val(ws, r, 7, self.meta.get(key2, ""), h="left", colspan=4)
            r += 1

        # ── Section 3: Calibration Setup ─────────────────────────────────────
        _blank_row(ws, r, N); r += 1
        _sec(ws, r, 1, "3. 교정 설정 (Calibration Setup)", colspan=N); r += 1

        setup_pairs = [
            ("센서 타입 (Sensor)",     self.sensor.name,
             "케이블 (Cable)",         self.meta.get("cable", "전용케이블")),
            ("공칭 저항 (Nominal)",    f"{self.sensor.r_nominal:.0f} Ω",
             "Inst. Amp. Gain",        str(self.meta.get("inst_amp_gain", "1"))),
            ("샘플링 속도",             f"{self.meta.get('sampling_hz', 100)} Hz",
             "측정 시간",              f"{self.meta.get('duration_sec', '-')} 초"),
            ("모사저항 값",
             " / ".join(f"{r:.0f}Ω" for r in self.resistances),
             "허용 편차 (Tolerance)",
             f"±{self.sensor.tolerance_ohm:.3f} Ω"),
        ]
        for lbl1, v1, lbl2, v2 in setup_pairs:
            _lbl(ws, r, 1, lbl1, bold=True)
            _val(ws, r, 2, v1, h="left", colspan=3)
            _lbl(ws, r, 6, lbl2, bold=True)
            _val(ws, r, 7, v2, h="left", colspan=4)
            r += 1

        # ── Section 4: Results Summary 2-1 ───────────────────────────────────
        _blank_row(ws, r, N); r += 1
        _sec(ws, r, 1,
             "4. 교정 결과 요약 — Method 2-1  (R_nom 기준 Offset)",
             colspan=N); r += 1

        r = self._result_table(ws, r, method="2-1")

        # ── Section 5: Results Summary 2-2 ───────────────────────────────────
        _blank_row(ws, r, N); r += 1
        _sec(ws, r, 1,
             "5. 교정 결과 요약 — Method 2-2  (편차 평균 Offset)",
             colspan=N); r += 1

        r = self._result_table(ws, r, method="2-2")

        # ── Signature block ───────────────────────────────────────────────────
        _blank_row(ws, r, N); r += 1
        sig_labels = ["작성 (Prepared)", "검토 (Reviewed)", "승인 (Approved)"]
        for i, label in enumerate(sig_labels):
            col = i * 3 + 1
            ws.merge_cells(start_row=r,   start_column=col, end_row=r,   end_column=col+2)
            ws.merge_cells(start_row=r+1, start_column=col, end_row=r+2, end_column=col+2)
            lc = ws.cell(row=r,   column=col, value=label)
            lc.fill      = _fill(C_LBLUE)
            lc.font      = _font(bold=True, size=8)
            lc.alignment = _align()
            lc.border    = _thin_border()
            bc = ws.cell(row=r+1, column=col, value="")
            bc.border    = _thin_border()
            ws.row_dimensions[r+1].height = 30
            ws.row_dimensions[r+2].height = 6

        # ── Column widths ────────────────────────────────────────────────────
        col_widths = [16, 14, 10, 10, 4, 16, 14, 10, 10, 4]
        for ci, w in enumerate(col_widths, 1):
            ws.column_dimensions[get_column_letter(ci)].width = w

    def _result_table(self, ws, start_row: int, method: str) -> int:
        """Write a 16-channel results table; return next free row."""
        rs   = self.resistances
        n_rs = len(rs)
        tol  = self.sensor.tolerance_ohm

        # Header row 1
        r = start_row
        _hdr(ws, r, 1, "채널")
        _hdr(ws, r, 2, "Excitation")
        _hdr(ws, r, 3, "Gain")
        _hdr(ws, r, 4, "Offset")
        for j, res in enumerate(rs):
            _hdr(ws, r, 5 + j, f"{res:.0f}Ω 편차")
        _hdr(ws, r, 5 + n_rs, "판정")
        r += 1

        # Header row 2 (units)
        _hdr(ws, r, 1, "-", size=7)
        _hdr(ws, r, 2, "V/Ω", size=7)
        _hdr(ws, r, 3, "-", size=7)
        _hdr(ws, r, 4, "Ω", size=7)
        for j in range(n_rs):
            _hdr(ws, r, 5 + j, "Ω", size=7)
        _hdr(ws, r, 5 + n_rs, "-", size=7)
        r += 1

        # Data rows
        for ri, ch in enumerate(self.channels):
            cal = self.cals[ch]
            alt = _fill(C_ALTROW) if ri % 2 == 0 else None

            def _dv(col, value, fmt=None, fc=C_BLACK):
                c = ws.cell(row=r, column=col, value=value)
                c.font      = _font(color=fc, size=8)
                c.alignment = _align()
                c.border    = _thin_border()
                if fmt:
                    c.number_format = fmt
                if alt:
                    c.fill = alt

            if method == "2-1":
                offset = cal.offset_100
                devs   = {res: cal.dev_final_100.get(res) for res in rs}
            else:
                offset = cal.offset_mean
                devs   = {res: cal.dev_final_mean.get(res) for res in rs}

            _dv(1, ch)
            _dv(2, cal.excitation, "0.0000")
            _dv(3, cal.gain,       "0.000000")
            _dv(4, offset,         "0.000000")

            pass_all = True
            for j, res in enumerate(rs):
                dev = devs.get(res)
                if dev is None:
                    _dv(5 + j, "-")
                else:
                    ok = abs(dev) <= tol
                    if not ok:
                        pass_all = False
                    _dv(5 + j, dev, "0.0000", C_PASS if ok else C_FAIL)

            _dv(5 + n_rs, "PASS" if pass_all else "FAIL",
                fc=C_PASS if pass_all else C_FAIL)
            r += 1

        return r

    # ── Per-channel sheets ────────────────────────────────────────────────────

    def _build_channel(self, wb: Workbook, ch_idx: int, ch: str):
        ws = wb.create_sheet(ch)
        ws.sheet_view.showGridLines = False

        cal = self.cals[ch]
        rs  = self.resistances
        n   = len(rs)
        tol = self.sensor.tolerance_ohm

        # Column layout: A=section, B=item, C…=resistance values
        COL_SEC  = 1
        COL_ITEM = 2
        COL_RES  = 3   # first resistance column

        ws.column_dimensions["A"].width = 22
        ws.column_dimensions["B"].width = 14
        for j in range(n):
            ws.column_dimensions[get_column_letter(COL_RES + j)].width = 13

        # ── Page title ───────────────────────────────────────────────────────
        total = len(self.channels)
        ws.merge_cells(start_row=1, start_column=1,
                       end_row=1, end_column=2 + n)
        t = ws.cell(row=1, column=1,
                    value=f"Calibration Sheet ({ch_idx+1}/{total})  —  {ch}")
        t.font      = Font(bold=True, size=12, color=C_NAVY, name="맑은 고딕")
        t.alignment = _align(h="left", v="center")
        ws.row_dimensions[1].height = 22

        # ── Result summary (2-1 and 2-2) ─────────────────────────────────────
        r = 3
        _sec(ws, r, COL_SEC, "교정 결과 요약", colspan=2 + n); r += 1

        # Header
        _hdr(ws, r, COL_SEC,  "구분")
        _hdr(ws, r, COL_ITEM, "채널명")
        # Will add more header cols below; using inline approach
        header_extra = ["Excitation", "Gain", "Offset"] + [f"{x:.0f}Ω" for x in rs]
        for j, txt in enumerate(header_extra):
            _hdr(ws, r, COL_RES + j, txt)
        r += 1

        def _res_row(label, offset_val, devs_dict):
            _sec(ws, r, COL_SEC, label)
            _val(ws, r, COL_ITEM, ch)
            _val(ws, r, COL_RES,     cal.excitation, "0.0000")
            _val(ws, r, COL_RES + 1, cal.gain,       "0.000000")
            _val(ws, r, COL_RES + 2, offset_val,     "0.000000")
            for j, res in enumerate(rs):
                dev = devs_dict.get(res)
                if dev is None:
                    _val(ws, r, COL_RES + 3 + j, "-")
                else:
                    _val(ws, r, COL_RES + 3 + j, dev, "0.0000",
                         color=_pass_color(dev, tol))

        # Re-map columns for summary: Excitation, Gain, Offset, then rs
        # Overwrite header to match
        prev_r = r - 1
        hcols = ["Excitation", "Gain", "Offset"] + [f"{x:.0f}Ω" for x in rs]
        for j, txt in enumerate(hcols):
            ws.cell(row=prev_r, column=COL_RES + j).value = txt

        _res_row("2-1  (R_nom offset)", cal.offset_100,  cal.dev_final_100)
        r += 1
        _res_row("2-2  (Mean offset)",  cal.offset_mean, cal.dev_final_mean)
        r += 1

        # ── Charts ───────────────────────────────────────────────────────────
        _blank_row(ws, r, 2 + n); r += 1
        _sec(ws, r, COL_SEC, "그래프", colspan=2 + n); r += 1

        gain_buf = _chart_gain(cal)
        dev_buf  = _chart_deviation(cal, rs)
        chart_row = r

        img_g = XLImage(gain_buf)
        img_g.width  = 340
        img_g.height = 210
        ws.add_image(img_g, ws.cell(row=chart_row, column=COL_SEC).coordinate)

        img_d = XLImage(dev_buf)
        img_d.width  = 340
        img_d.height = 210
        ws.add_image(img_d, ws.cell(row=chart_row, column=COL_RES).coordinate)

        r += 14   # approximate rows occupied by charts at default row height

        # ── Detail section ────────────────────────────────────────────────────
        _blank_row(ws, r, 2 + n); r += 1
        _sec(ws, r, COL_SEC, "상세 정보 (Detail)", colspan=2 + n); r += 1

        # Sub-header
        _hdr(ws, r, COL_SEC,  "구분")
        _hdr(ws, r, COL_ITEM, "항목")
        for j, res in enumerate(rs):
            _hdr(ws, r, COL_RES + j, f"{res:.0f}Ω")
        r += 1

        def _drow(section_label, item_label, values,
                  sec_fill=None, alt=False, fmt="0.0000", color=C_BLACK):
            sc = ws.cell(row=r, column=COL_SEC,  value=section_label)
            ic = ws.cell(row=r, column=COL_ITEM, value=item_label)
            for c in (sc, ic):
                c.font      = _font(bold=(section_label != ""), size=8)
                c.alignment = _align(h="left")
                c.border    = _thin_border()
                if sec_fill:
                    c.fill = _fill(sec_fill)
                elif alt:
                    c.fill = _fill(C_ALTROW)

            for j, v in enumerate(values):
                vc = ws.cell(row=r, column=COL_RES + j, value=v)
                vc.font      = _font(color=color, size=8)
                vc.alignment = _align()
                vc.border    = _thin_border()
                if v is not None and isinstance(v, float):
                    vc.number_format = fmt
                if sec_fill:
                    vc.fill = _fill(sec_fill)
                elif alt:
                    vc.fill = _fill(C_ALTROW)

        # 1) Decimal
        _drow("1) Decimal", "AVG",
              [int(cal.decimals_avg.get(x, 0)) for x in rs],
              sec_fill=C_LBLUE, fmt="0")
        r += 1
        _drow("", "MIN",
              [int(cal.decimals_min.get(x, 0)) for x in rs], fmt="0")
        r += 1
        _drow("", "MAX",
              [int(cal.decimals_max.get(x, 0)) for x in rs],
              alt=True, fmt="0")
        r += 1

        # 2) Voltage before gain
        _drow("2) 전압 (before gain)", "Ref",
              [cal.voltage_ref.get(x) for x in rs], sec_fill=C_LBLUE)
        r += 1
        _drow("", "AVG",
              [cal.voltages_avg.get(x) for x in rs])
        r += 1
        _drow("", "MIN",
              [cal.voltages_min.get(x) for x in rs], alt=True)
        r += 1
        _drow("", "MAX",
              [cal.voltages_max.get(x) for x in rs])
        r += 1
        _drow("", "편차 (AVG-Ref)",
              [cal.voltages_avg.get(x, 0) - cal.voltage_ref.get(x, 0)
               if x in cal.voltages_avg else None for x in rs],
              alt=True)
        r += 1

        # 3) Resistance before gain
        _drow("3) 저항 (before gain)", "Ref",
              [x for x in rs], sec_fill=C_LBLUE)
        r += 1
        _drow("", "AVG",
              [cal.r_before_gain.get(x) for x in rs])
        r += 1
        _drow("", "편차",
              [cal.dev_before_gain.get(x) for x in rs], alt=True)
        r += 1

        # 4) Gain
        gc = ws.cell(row=r, column=COL_SEC,  value="4) Gain")
        gv = ws.cell(row=r, column=COL_ITEM, value=f"{cal.gain:.8f}")
        for c in (gc, gv):
            c.fill      = _fill(C_LBLUE)
            c.font      = _font(bold=True, size=8)
            c.alignment = _align(h="left")
            c.border    = _thin_border()
        ws.merge_cells(start_row=r, start_column=COL_RES,
                       end_row=r, end_column=COL_RES + n - 1)
        ws.cell(row=r, column=COL_RES).border = _thin_border()
        r += 1

        # 5) Resistance after gain
        _drow("5) 저항 (after gain)", "AVG",
              [cal.r_after_gain_avg.get(x) for x in rs], sec_fill=C_LBLUE)
        r += 1
        _drow("", "MIN",
              [cal.r_after_gain_min.get(x) for x in rs])
        r += 1
        _drow("", "MAX",
              [cal.r_after_gain_max.get(x) for x in rs], alt=True)
        r += 1
        _drow("", "편차",
              [cal.dev_after_gain.get(x) for x in rs])
        r += 1

        # 6) Offsets
        oc = ws.cell(row=r, column=COL_SEC,  value="6) Offset")
        ov = ws.cell(row=r, column=COL_ITEM,
                     value=f"2-1: {cal.offset_100:.6f}  /  2-2: {cal.offset_mean:.6f}")
        for c in (oc, ov):
            c.fill      = _fill(C_LBLUE)
            c.font      = _font(bold=True, size=8)
            c.alignment = _align(h="left")
            c.border    = _thin_border()
        ws.merge_cells(start_row=r, start_column=COL_RES,
                       end_row=r, end_column=COL_RES + n - 1)
        ws.cell(row=r, column=COL_RES).border = _thin_border()
        r += 1

        # 7) Final 2-1
        _drow("7) 2-1 최종 (R_nom offset)", "AVG",
              [cal.r_final_100_avg.get(x) for x in rs], sec_fill=C_LBLUE)
        r += 1
        devs_100 = [cal.dev_final_100.get(x) for x in rs]
        colors_100 = [(_pass_color(d, tol) if d is not None else C_BLACK)
                      for d in devs_100]
        # write deviation with per-cell coloring
        sc = ws.cell(row=r, column=COL_SEC,  value="")
        ic = ws.cell(row=r, column=COL_ITEM, value="편차")
        for c in (sc, ic):
            c.font = _font(size=8); c.alignment = _align(h="left"); c.border = _thin_border()
        for j, (d, fc) in enumerate(zip(devs_100, colors_100)):
            vc = ws.cell(row=r, column=COL_RES + j, value=d)
            vc.font = _font(color=fc, size=8); vc.alignment = _align()
            vc.border = _thin_border()
            if isinstance(d, float): vc.number_format = "0.0000"
        r += 1
        _drow("", "허용 (+)",
              [cal.tolerance_max_100.get(x, tol) for x in rs], alt=True)
        r += 1
        _drow("", "허용 (-)",
              [-cal.tolerance_max_100.get(x, tol) for x in rs])
        r += 1

        # 8) Final 2-2
        _drow("8) 2-2 최종 (Mean offset)", "AVG",
              [cal.r_final_mean_avg.get(x) for x in rs], sec_fill=C_LBLUE)
        r += 1
        devs_mean = [cal.dev_final_mean.get(x) for x in rs]
        colors_mean = [(_pass_color(d, tol) if d is not None else C_BLACK)
                       for d in devs_mean]
        sc = ws.cell(row=r, column=COL_SEC,  value="")
        ic = ws.cell(row=r, column=COL_ITEM, value="편차")
        for c in (sc, ic):
            c.font = _font(size=8); c.alignment = _align(h="left"); c.border = _thin_border()
        for j, (d, fc) in enumerate(zip(devs_mean, colors_mean)):
            vc = ws.cell(row=r, column=COL_RES + j, value=d)
            vc.font = _font(color=fc, size=8); vc.alignment = _align()
            vc.border = _thin_border()
            if isinstance(d, float): vc.number_format = "0.0000"
        r += 1
        _drow("", "허용 (+)",
              [cal.tolerance_max_mean.get(x, tol) for x in rs], alt=True)
        r += 1
        _drow("", "허용 (-)",
              [-cal.tolerance_max_mean.get(x, tol) for x in rs])

    # ── 3-Wire Reference sheet ────────────────────────────────────────────────

    def _build_3w_reference(self, wb: Workbook):
        ws = wb.create_sheet("3W_Reference")
        ws.sheet_view.showGridLines = False

        gains = [g for g in STANDARD_GAINS if g in (1, 10, 100)]
        sensor_type = self.sensor.sensor_type

        # Determine resistance range for the sensor
        r_nom = self.sensor.r_nominal
        r_start = max(0, r_nom * 0.5)
        r_end   = r_nom * 1.5
        r_step  = 1.0

        try:
            resistances, gain_list, table = build_reference_table(
                sensor_type, r_start, r_end, r_step, gains
            )
        except Exception:
            ws["A1"].value = "3-Wire 레퍼런스 테이블 생성 실패"
            return

        # Title
        ws.merge_cells(f"A1:{get_column_letter(3 + len(gains))}1")
        t = ws["A1"]
        t.value     = f"3-Wire Reference Voltage  [{self.sensor.name}]"
        t.font      = Font(bold=True, size=12, color=C_NAVY, name="맑은 고딕")
        t.alignment = _align(h="left", v="center")
        ws.row_dimensions[1].height = 22

        # Header
        r = 3
        _hdr(ws, r, 1, f"R (Ω)")
        _hdr(ws, r, 2, f"ΔR (Ω)")
        for j, g in enumerate(gain_list):
            _hdr(ws, r, 3 + j, f"Gain={g:.0f}")
        r += 1

        # Data
        for ri, res in enumerate(resistances):
            delta = res - r_nom
            alt   = _fill(C_ALTROW) if ri % 2 == 0 else None
            for col_idx, value in enumerate([res, delta] + list(table[res])):
                c = ws.cell(row=r, column=1 + col_idx, value=value)
                c.font      = _font(size=8)
                c.alignment = _align()
                c.border    = _thin_border()
                if col_idx >= 2:
                    c.number_format = "0.000000"
                if alt:
                    c.fill = alt
            r += 1

        # Column widths
        ws.column_dimensions["A"].width = 10
        ws.column_dimensions["B"].width = 10
        for j in range(len(gain_list)):
            ws.column_dimensions[get_column_letter(3 + j)].width = 14

    # ── Raw Data sheet ────────────────────────────────────────────────────────

    def _build_raw_data(self, wb: Workbook):
        ws = wb.create_sheet("Raw_Data")
        ws.sheet_view.showGridLines = False

        rs = self.resistances
        chs = self.channels

        # Title
        ws.merge_cells(f"A1:{get_column_letter(2 + len(chs))}1")
        t = ws["A1"]
        t.value     = "Raw Decimal Data  (AVG / MIN / MAX per resistance step)"
        t.font      = Font(bold=True, size=11, color=C_NAVY, name="맑은 고딕")
        t.alignment = _align(h="left", v="center")
        ws.row_dimensions[1].height = 20

        # Header
        r = 3
        _hdr(ws, r, 1, "모사저항 (Ω)")
        _hdr(ws, r, 2, "항목")
        for j, ch in enumerate(chs):
            _hdr(ws, r, 3 + j, ch)
        r += 1

        for ri, res in enumerate(rs):
            for si, stat_key in enumerate(["AVG", "MIN", "MAX"]):
                alt = _fill(C_ALTROW) if ri % 2 == 0 else None
                lbl_val = f"{res:.0f}" if si == 0 else ""
                lc = ws.cell(row=r, column=1, value=lbl_val)
                sc = ws.cell(row=r, column=2, value=stat_key)
                for c in (lc, sc):
                    c.font = _font(size=8); c.alignment = _align(); c.border = _thin_border()
                    if alt: c.fill = alt

                for j, ch in enumerate(chs):
                    cal = self.cals.get(ch)
                    if cal is None:
                        v = None
                    elif stat_key == "AVG":
                        v = int(cal.decimals_avg.get(res, 0))
                    elif stat_key == "MIN":
                        v = int(cal.decimals_min.get(res, 0))
                    else:
                        v = int(cal.decimals_max.get(res, 0))
                    vc = ws.cell(row=r, column=3 + j, value=v)
                    vc.font = _font(size=8); vc.alignment = _align(); vc.border = _thin_border()
                    if alt: vc.fill = alt
                r += 1

        # Column widths
        ws.column_dimensions["A"].width = 14
        ws.column_dimensions["B"].width = 8
        for j in range(len(chs)):
            ws.column_dimensions[get_column_letter(3 + j)].width = 9

    # ── Public API ────────────────────────────────────────────────────────────

    def write(self, output_path: str) -> str:
        os.makedirs(os.path.dirname(output_path), exist_ok=True)
        wb = Workbook()

        self._build_cover(wb)
        for idx, ch in enumerate(self.channels):
            self._build_channel(wb, idx, ch)
        self._build_3w_reference(wb)
        self._build_raw_data(wb)

        wb.save(output_path)
        return output_path
